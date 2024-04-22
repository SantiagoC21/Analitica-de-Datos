from webbrowser import Mozilla
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook, Workbook

def datos_Noticias_RPP(url):
    d = {}
    headers = {"user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"}
    print(f'Realizando la petición: {url}')
    req = requests.get(url, headers=headers, timeout=10)
    print(f'Código de respuesta... {req.status_code} {req.reason}')
    if req.status_code != 200:
        return {"ERROR" : f"{req.reason}", "status_code":f"{req.status_code}"}
    d["url"] = req.url
    d["id"] = d["url"].split("-")[-1]
    soup = BeautifulSoup(req.text, 'html.parser')
    d["CATEGORIA"] = soup.find('ul', class_='breadcrump').text.split()[0]
    d["TITULO"] = soup.find('h1', class_='article__title').text.strip()
    d["FECHA"] = soup.find('div', class_="article__author").find('div').find('time').find('a').text
    d["AUTOR"] = soup.find('div', class_="article__author").find('div').find('p').find('a').text
    d["DIARIO"] = "RPP"
    d["IMAGEN"] = soup.find('div', class_='column-fluid').find('figure', class_='media').find('img').attrs.get('src')
    return d

def guardar_excel(datos, filename="C:/Users/piter/OneDrive/Escritorio/UNI/6TO CICLO/ANALITICA DE DATOS/ChatBot/ChatBot_DataBase.xlsx"):
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.append(list(datos.keys()))
        
    hoja = wb.active
    hoja.append(list(datos.values()))
    wb.save(filename)

if __name__ == '__main__':
    while True:
        URL = input('DIGITE URL O EXIT PARA SALIR: ')
        if URL.lower() == 'exit':
            break
        datos = datos_Noticias_RPP(URL)
        guardar_excel(datos)
        


    
 




 
    